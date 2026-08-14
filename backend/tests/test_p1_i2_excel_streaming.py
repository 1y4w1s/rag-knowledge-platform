"""M8-I2: xlsx 逐行迭代流式解析（防 list() 物化整表 OOM）。"""

from __future__ import annotations

from collections.abc import Iterator
from pathlib import Path

import openpyxl

from app.services.ingestion.parser import _rows_to_markdown_table, parse_xlsx


def _make_two_sheet_xlsx(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "部门预算"
    ws1.append(["部门", "金额", "备注"])
    ws1.append(["研发部", 500, None])
    ws1.append(["市场部", 0, False])
    ws1.append(["行政部", 0.0, ""])

    ws2 = wb.create_sheet("项目支出")
    ws2.append(["项目", "负责人"])
    ws2.append(["卧铺 v1", "张三"])
    wb.save(str(path))


def test_parse_xlsx_streaming_result_equivalent(tmp_path: Path) -> None:
    path = tmp_path / "two_sheets.xlsx"
    _make_two_sheet_xlsx(path)

    blocks = parse_xlsx(path)

    assert len(blocks) == 2
    assert [b.section_title for b in blocks] == ["部门预算", "项目支出"]
    assert [b.heading_path for b in blocks] == ["部门预算", "项目支出"]
    assert all(b.block_kind == "table" for b in blocks)
    assert blocks[0].content == (
        "| 部门 | 金额 | 备注 |\n"
        "| --- | --- | --- |\n"
        "| 研发部 | 500 |  |\n"
        "| 市场部 |  |  |\n"
        "| 行政部 |  |  |"
    )
    assert blocks[1].content == (
        "| 项目 | 负责人 |\n"
        "| --- | --- |\n"
        "| 卧铺 v1 | 张三 |"
    )


def test_rows_to_markdown_table_consumes_lazy_generator() -> None:
    consumed = 0

    def rows() -> Iterator[tuple]:
        nonlocal consumed
        for i in range(1, 6):
            consumed += 1
            yield (f"r{i}", i)

    gen = rows()
    assert consumed == 0  # 惰性生成器不应被 helper 预取成 list
    lines = _rows_to_markdown_table(gen)

    assert lines is not None
    assert consumed == 5
    assert lines == [
        "| r1 | 1 |",
        "| --- | --- |",
        "| r2 | 2 |",
        "| r3 | 3 |",
        "| r4 | 4 |",
        "| r5 | 5 |",
    ]


def test_parser_source_has_no_list_materialization() -> None:
    parser_path = Path(parse_xlsx.__code__.co_filename)
    source = parser_path.read_text(encoding="utf-8")
    assert "list(ws.iter_rows" not in source
    assert source.count("iter_rows(values_only=True)") == 1


def test_empty_sheet_returns_no_blocks(tmp_path: Path) -> None:
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    wb.save(str(path))

    assert parse_xlsx(path) == []


def test_first_row_all_none_skipped(tmp_path: Path) -> None:
    path = tmp_path / "blank_header.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None])
    ws.append(["alpha", 1])
    wb.save(str(path))

    assert parse_xlsx(path) == []


def test_empty_sheet_before_normal_sheet_read_only(tmp_path: Path) -> None:
    path = tmp_path / "mixed.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "空表"
    ws2 = wb.create_sheet("正常表")
    ws2.append(["名称", "数量"])
    ws2.append(["电池", 3])
    wb.save(str(path))

    blocks = parse_xlsx(path)
    assert len(blocks) == 1
    assert blocks[0].section_title == "正常表"
    assert blocks[0].content == "| 名称 | 数量 |\n| --- | --- |\n| 电池 | 3 |"


def test_falsy_cells_render_empty_strings() -> None:
    rows: list[tuple] = [("v",), (0,), (False,), (0.0,), (None,), ("",)]
    lines = _rows_to_markdown_table(iter(rows))

    assert lines == [
        "| v |",
        "| --- |",
        "|  |",
        "|  |",
        "|  |",
        "|  |",
        "|  |",
    ]


def test_large_generator_fully_consumed() -> None:
    total = 20_000
    consumed = 0

    def rows() -> Iterator[tuple]:
        nonlocal consumed
        yield ("row", "h")
        for i in range(total):
            consumed += 1
            yield ("row", i)

    lines = _rows_to_markdown_table(rows())
    assert lines is not None
    assert consumed == total
    assert len(lines) == total + 2  # header + separator + data rows
    assert lines[0] == "| row | h |"
    assert lines[-1] == f"| row | {total - 1} |"
